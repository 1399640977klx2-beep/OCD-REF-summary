"""Second pivot: Product/Wafer rows, Tool columns, per-parameter stats.

Standalone script; it does not import other modules in this project.

Run:
    python utils/pivot_tool_summary.py                          # 使用默认 Test1.xlsx
    python utils/pivot_tool_summary.py <input.xlsx> <output.xlsx>
    python utils/pivot_tool_summary.py <input.xlsx> <output.xlsx> --sheet Sheet1

Layout (all parameters in one sheet):
    Row 1: parameter name, merged across each parameter block
    Row 2: Product, Wafer, <Tool>_PMISH_Mean ..., REF_Mean,
           <Tool>_Bias_Mean ..., <Tool>_PMISH_STD ..., REF_STD
    Row 3+: one row per Product + Wafer; Product/Wafer repeat only when changed.

Bias logic: prefer <param>_bias_Addoffset, fall back to <param>_Bias.
"""
import argparse
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

DEFAULT_PARAMETERS = ["CD_Bot", "CD_Top", "HIGH", "SPA", "THK"]

DEFAULT_INPUT_PATH = (
    Path(__file__).resolve().parents[1] / "Data" / "绘制数据透视表" / "Test1.xlsx"
)


def _norm(value):
    return str(value).strip().lower()


def _find_header_row(raw):
    """Return the 0-based row index of the real data header."""
    limit = min(len(raw), 200)
    for idx in range(limit):
        values = {_norm(v) for v in raw.iloc[idx].tolist()}
        if "die seq" in values and "wafer id" in values:
            return idx
    for idx in range(limit):
        values = {_norm(v) for v in raw.iloc[idx].tolist()}
        if "wafer id" in values:
            return idx
    raise ValueError("未找到数据表头行：需要包含 Die Seq 和 Wafer ID")


def _read_raw(path, sheet=None):
    path = Path(path)
    if path.suffix.lower() == ".csv":
        return pd.read_csv(path, header=None, dtype=str)
    return pd.read_excel(
        path, sheet_name=sheet if sheet is not None else 0,
        header=None, dtype=str,
    )


def load_dataframe(path, sheet=None):
    """Read pasted data and return a DataFrame with the real header applied."""
    raw = _read_raw(path, sheet)
    header_idx = _find_header_row(raw)
    header = [
        str(v).strip() if not pd.isna(v) else f"Unnamed_{i}"
        for i, v in enumerate(raw.iloc[header_idx])
    ]
    df = raw.iloc[header_idx + 1:].copy()
    df.columns = header
    df = df.reset_index(drop=True)
    if df.empty:
        raise ValueError("表头下方没有数据")
    return df


def _find_group_columns(df):
    norm_to_orig = {}
    for col in df.columns:
        norm_to_orig.setdefault(_norm(col), col)

    tool_col = norm_to_orig.get("tool")
    product_col = norm_to_orig.get("product")
    wafer_col = norm_to_orig.get("wafer id") or norm_to_orig.get("waferid")

    missing = []
    if tool_col is None:
        missing.append("Tool")
    if product_col is None:
        missing.append("Product")
    if wafer_col is None:
        missing.append("Wafer ID")
    if missing:
        raise ValueError("缺少分组列: " + ", ".join(missing))
    return tool_col, product_col, wafer_col


def _classify_suffix(suffix):
    compact = "".join(ch for ch in _norm(suffix) if ch.isalnum())
    if compact == "pmish":
        return "pmish"
    if compact == "ref":
        return "ref"
    if compact == "bias":
        return "bias"
    if "addoffset" in compact:
        return "bias_adoffset"
    return None


def _collect_param_columns(df, parameters):
    found = {
        p: {"pmish": None, "ref": None, "bias": None, "bias_adoffset": None}
        for p in parameters
    }
    for col in df.columns:
        norm = _norm(col)
        for p in parameters:
            prefix = _norm(p) + "_"
            if norm.startswith(prefix):
                kind = _classify_suffix(norm[len(prefix):])
                if kind:
                    found[p][kind] = col
    return found


def _resolve_param_columns(found, parameters):
    warnings = []
    resolved = {}
    for p in parameters:
        m = found[p]
        if m["pmish"] is None:
            warnings.append(f"缺少 {p}_PMISH 后缀")
        if m["ref"] is None:
            warnings.append(f"缺少 {p}_REF 后缀")

        bias_col = m["bias_adoffset"]
        if bias_col is None and m["bias"] is not None:
            bias_col = m["bias"]
            warnings.append(f"缺少 {p}_bias_Addoffset 后缀，改用 {m['bias']}")
        if bias_col is None:
            warnings.append(f"缺少 {p} 的 Bias 后缀（bias_Addoffset 或 Bias）")

        resolved[p] = {
            "pmish": m["pmish"],
            "ref": m["ref"],
            "bias": bias_col,
        }
    return resolved, warnings


def _display(value):
    try:
        if value is None or pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    return value


def _key_text(value):
    return str(_display(value))


def _same_key(a, b):
    return _display(a) == _display(b)


def _to_numeric(series):
    return pd.to_numeric(series, errors="coerce")


def _empty_if_unavailable(value):
    try:
        if value is None or pd.isna(value):
            return ""
    except (TypeError, ValueError):
        return ""
    return value


def _agg_map(df, group_cols, value_col, stat):
    if value_col is None or value_col not in df.columns:
        return {}
    values = _to_numeric(df[value_col])
    grouped = values.groupby(
        [df[c] for c in group_cols], dropna=False, sort=False)
    agg = grouped.mean() if stat == "mean" else grouped.std()
    result = {}
    for keys, value in agg.items():
        key = tuple(keys) if isinstance(keys, tuple) else (keys,)
        result[tuple(_key_text(k) for k in key)] = _empty_if_unavailable(value)
    return result


def _build_column_defs(parameters, tools):
    column_defs = []
    blocks = []
    for idx, p in enumerate(parameters):
        block_defs = []
        for t in tools:
            block_defs.append((p, "pmish_mean", t, f"{t}_PMISH_Mean"))
        block_defs.append((p, "ref_mean", None, "REF_Mean"))
        for t in tools:
            block_defs.append((p, "bias_mean", t, f"{t}_Bias_Mean"))
        for t in tools:
            block_defs.append((p, "pmish_std", t, f"{t}_PMISH_STD"))
        block_defs.append((p, "ref_std", None, "REF_STD"))

        column_defs.extend(block_defs)
        blocks.append((p, block_defs))
        if idx < len(parameters) - 1:
            column_defs.append((None, "blank", None, ""))
    return column_defs, blocks


def compute_tool_pivot(df, parameters=None):
    if parameters is None:
        parameters = DEFAULT_PARAMETERS

    tool_col, product_col, wafer_col = _find_group_columns(df)
    found = _collect_param_columns(df, parameters)
    resolved, warnings = _resolve_param_columns(found, parameters)

    tools = sorted(
        {_key_text(v) for v in df[tool_col].dropna().unique()},
        key=str.lower,
    )
    if not tools:
        raise ValueError("未找到任何 Tool 数据")

    pmish_mean = {}
    pmish_std = {}
    bias_mean = {}
    ref_mean = {}
    ref_std = {}

    for p in parameters:
        r = resolved[p]
        for (prod, wafer, tool), value in _agg_map(
            df, [product_col, wafer_col, tool_col], r["pmish"], "mean"
        ).items():
            pmish_mean[(p, prod, wafer, tool)] = value
        for (prod, wafer, tool), value in _agg_map(
            df, [product_col, wafer_col, tool_col], r["pmish"], "std"
        ).items():
            pmish_std[(p, prod, wafer, tool)] = value
        for (prod, wafer, tool), value in _agg_map(
            df, [product_col, wafer_col, tool_col], r["bias"], "mean"
        ).items():
            bias_mean[(p, prod, wafer, tool)] = value
        for (prod, wafer), value in _agg_map(
            df, [product_col, wafer_col], r["ref"], "mean"
        ).items():
            ref_mean[(p, prod, wafer)] = value
        for (prod, wafer), value in _agg_map(
            df, [product_col, wafer_col], r["ref"], "std"
        ).items():
            ref_std[(p, prod, wafer)] = value

    pairs = sorted(
        {
            (_key_text(prod), _key_text(wafer))
            for prod, wafer in df[[product_col, wafer_col]].itertuples(index=False)
        },
        key=lambda x: (x[0], x[1]),
    )

    column_defs, blocks = _build_column_defs(parameters, tools)
    column_names = ["Product", "Wafer"] + [name for *_, name in column_defs]

    rows = []
    prev_product = prev_wafer = None
    for prod, wafer in pairs:
        row_values = [
            "" if _same_key(prod, prev_product) else prod,
            "" if _same_key(wafer, prev_wafer) else wafer,
        ]
        for p, kind, tool, _name in column_defs:
            if kind == "blank":
                row_values.append("")
            elif kind == "pmish_mean":
                row_values.append(pmish_mean.get((p, prod, wafer, tool), ""))
            elif kind == "ref_mean":
                row_values.append(ref_mean.get((p, prod, wafer), ""))
            elif kind == "bias_mean":
                row_values.append(bias_mean.get((p, prod, wafer, tool), ""))
            elif kind == "pmish_std":
                row_values.append(pmish_std.get((p, prod, wafer, tool), ""))
            elif kind == "ref_std":
                row_values.append(ref_std.get((p, prod, wafer), ""))
        rows.append(row_values)
        prev_product, prev_wafer = prod, wafer

    return column_names, rows, blocks, warnings


def save_tool_pivot(output_path, column_names, rows, blocks, warnings=None):
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    # Row 1: merged parameter names.
    col = 3
    for p, block_defs in blocks:
        end = col + len(block_defs) - 1
        if end > col:
            ws.merge_cells(start_row=1, start_column=col,
                           end_row=1, end_column=end)
        cell = ws.cell(row=1, column=col, value=p)
        cell.font = header_font
        cell.alignment = center
        col = end + 2

    # Row 2: column names.
    for j, name in enumerate(column_names, 1):
        cell = ws.cell(row=2, column=j, value=name if name else None)
        cell.font = header_font
        cell.alignment = center

    # Data rows: leave missing cells untouched so they stay blank.
    for i, row_values in enumerate(rows, start=3):
        for j, value in enumerate(row_values, 1):
            if value != "":
                ws.cell(row=i, column=j, value=value)

    if warnings:
        ww = wb.create_sheet("Warnings")
        ww.append(["Warning"])
        for w in warnings:
            ww.append([w])

    wb.save(output_path)
    return output_path


def main(argv=None):
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "input", nargs="?", default=None,
        help="Excel/CSV 输入文件；不填时使用 Data\\绘制数据透视表\\Test1.xlsx",
    )
    parser.add_argument(
        "output", nargs="?", default=None,
        help="输出 xlsx 文件；不填时自动生成 <输入名>_tool_pivot.xlsx",
    )
    parser.add_argument("--sheet", default=None, help="Excel sheet 名称，默认取第一个 sheet")
    args = parser.parse_args(argv)

    input_path = args.input
    if input_path is None:
        input_path = str(DEFAULT_INPUT_PATH)
        if not Path(input_path).exists():
            parser.error(f"未提供输入文件，且默认文件不存在: {input_path}")
        print(f"[信息] 未提供输入文件，使用默认: {input_path}")

    df = load_dataframe(input_path, args.sheet)
    column_names, rows, blocks, warnings = compute_tool_pivot(df)
    for w in warnings:
        print(f"[警告] {w}", file=sys.stderr)

    output = args.output or str(
        Path(input_path).with_name(Path(input_path).stem + "_tool_pivot.xlsx"))
    save_tool_pivot(output, column_names, rows, blocks, warnings)
    print(f"已生成: {Path(output).resolve()}")
    print(f"行数: {len(rows)}（Product x Wafer），列数: {len(column_names)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
