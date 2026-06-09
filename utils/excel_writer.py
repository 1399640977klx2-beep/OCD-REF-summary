"""
Helper functions for writing Excel output files.
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)

HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=10)
DATA_FONT = Font(size=10)


def write_ref_summary_excel(df, output_path, vendor):
    """
    Write the REF summary Excel file.
    """
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Summary', index=False)
        ws = writer.sheets['Summary']
        _style_header(ws, df.columns)
        _auto_width(ws, df)


def write_match_excel(bsl_df, stats_df, match_df, output_path):
    """
    Write the Match Excel with 3 sheets.
    """
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        bsl_df.to_excel(writer, sheet_name='BSL', index=False)
        _style_header(writer.sheets['BSL'], bsl_df.columns)

        stats_df.to_excel(writer, sheet_name='Mean&STD', index=False)
        _style_header(writer.sheets['Mean&STD'], stats_df.columns)

        match_df.to_excel(writer, sheet_name='Match', index=False)
        _style_header(writer.sheets['Match'], match_df.columns)

        for ws_name in writer.sheets:
            ws = writer.sheets[ws_name]
            df = {'BSL': bsl_df, 'Mean&STD': stats_df, 'Match': match_df}[ws_name]
            _auto_width(ws, df)


def _style_header(ws, columns):
    for col_idx, _col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER


def _auto_width(ws, df):
    for col_idx, col_name in enumerate(df.columns, 1):
        max_len = max(
            df[col_name].astype(str).map(len).max() if len(df) > 0 else 0,
            len(str(col_name))
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 30)
def write_ref_summary_by_pad(df, output_path, group_col="Pad name"):
    """Write REF summary Excel with one sheet per PAD."""
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        pads = df[group_col].unique() if group_col in df.columns else ["Summary"]
        for pad in pads:
            subset = df[df[group_col] == pad] if group_col in df.columns else df
            subset.to_excel(writer, sheet_name=str(pad)[:31], index=False)
            _style_header(writer.sheets[str(pad)[:31]], subset.columns)
            _auto_width(writer.sheets[str(pad)[:31]], subset)
