"""
Match sheet generation for Tab 2.
"""

# Column name aliases for different BSL vendors
WAFER_COLS = ["WaferID", "Wafer ID", "Wafer"]
X_COLS = ["Col/X", "FIELD X", "RefDieCol"]
Y_COLS = ["Row/Y", "FIELD Y", "RefDieRow"]


def _detect_coord_cols(df):
    bsl_cols = set(df.columns)
    for x_cand in X_COLS:
        for y_cand in Y_COLS:
            if x_cand in bsl_cols and y_cand in bsl_cols:
                return x_cand, y_cand
    raise ValueError(
        "Cannot detect coordinate columns. "
        "Checked X: " + str(X_COLS) + ", Y: " + str(Y_COLS))


def _col_letter(n):
    # Convert 1-based column index to Excel letter (1->A, 28->AB)
    r = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        r = chr(65 + rem) + r
    return r


def _write_formula_col(ws, target_col, company_df, company_start, col_name, nrows, header, row_base=4):
    # Write header and formula references to a company data column
    if col_name not in company_df.columns:
        raise KeyError("Column " + repr(col_name) + " not found in company data")
    src_idx = company_df.columns.get_loc(col_name)
    letter = _col_letter(company_start + src_idx)
    ws.cell(row=3, column=target_col, value=header)
    for r in range(nrows):
        ws.cell(row=r + row_base, column=target_col, value="=" + letter + str(r + row_base))

def _write_calibrated_data_col(ws, target_col, nrows, header, numparam, row_base=4):
    ws.cell(row=3, column=target_col, value=f"{header}_Calibrated")
    for r in range(nrows):
        ws.cell(row=r + row_base, column=target_col, value=f"={_col_letter(target_col-numparam)}{r + row_base}"
                                                           f"*{_col_letter(target_col-numparam)}$1"
                                                           f"+{_col_letter(target_col-numparam)}$2")

def _write_ref_data_col(ws, target_col, nrows_company, nrows_bsl, header, numparam, currentparam, id_col, row_base=4):
    ws.cell(row=3, column=target_col, value=f"{header}_Ref")
    for r in range(nrows_company):
        ws.cell(row=r + row_base, column=target_col, value=f"=vlookup({_col_letter(id_col)}{r + row_base},"
                                                           f"$D$3:${_col_letter(numparam+4)}${nrows_bsl + row_base},"
                                                           f"{currentparam+2},"
                                                           f"FALSE)")

def _write_bias_data_col(ws, target_col, nrows_company, header, numparam, row_base=4):
    ws.cell(row=3, column=target_col, value=f"{header}_Bias")
    for r in range(nrows_company):
        cal = _col_letter(target_col - 2 * numparam)
        ref = _col_letter(target_col - numparam)
        ws.cell(row=r + row_base, column=target_col, value=f"={cal}{r + row_base}-{ref}{r + row_base}")

def _write_old_bias_data_col(ws, target_col, nrows_company, header, row_base=4):
    ws.cell(row=3, column=target_col, value=f"{header}_oldLibBias")
    for r in range(nrows_company):
        ws.cell(row=r + row_base, column=target_col,value=0)

def _safe_val(val):
    """Convert numpy/pandas values to Python native types."""
    import pandas as pd
    if pd.isna(val):
        return ""
    if hasattr(val, "item"):
        return val.item()
    return val


# Generate in a NEW file (copy BSL + company sheets)
def generate_match_sheet(src_path, dst_path, bsl_sheet, company_sheet, selected_params):
    import shutil
    from openpyxl import load_workbook
    import pandas as pd

    shutil.copy2(src_path, dst_path)

    wb = load_workbook(dst_path)
    if "Match" in wb.sheetnames:
        del wb["Match"]
    ws = wb.create_sheet("Match")

    bsl_df = pd.read_excel(dst_path, sheet_name=bsl_sheet)
    company_df = pd.read_excel(dst_path, sheet_name=company_sheet)

    n = len(selected_params)
    bsl_rows = len(bsl_df)
    company_rows = len(company_df)

    last_bsl_param_col = 4 + n
    company_start_col = last_bsl_param_col + (5 * n + 13)
    product_col = company_start_col - 1

    ws.cell(row=3, column=1, value="WaferID")
    coord_x, coord_y = _detect_coord_cols(bsl_df)
    ws.cell(row=3, column=2, value=coord_x)
    ws.cell(row=3, column=3, value=coord_y)
    ws.cell(row=3, column=4, value="ID")
    for idx, p in enumerate(selected_params):
        ws.cell(row=3, column=5 + idx, value=p)
    ws.cell(row=3, column=product_col, value="Product")
    for idx, col_name in enumerate(company_df.columns):
        ws.cell(row=3, column=company_start_col + idx, value=str(col_name)[:31])

    for r in range(bsl_rows):
        row = r + 4
        bd = bsl_df.iloc[r]
        ws.cell(row=row, column=1, value=str(bd.get("WaferID", bd.get("Wafer ID", ""))))
        ws.cell(row=row, column=2, value=_safe_val(bd.get(coord_x, "")))
        ws.cell(row=row, column=3, value=_safe_val(bd.get(coord_y, "")))
        ws.cell(row=row, column=4, value='=A%d&"_"&B%d&"_"&C%d' % (row, row, row))
        for p_idx, p in enumerate(selected_params):
            ws.cell(row=row, column=5 + p_idx, value=_safe_val(bd.get(p, "")))

    mid_start = last_bsl_param_col + 5

    # H, I, J: Hierarchical axis labels for charting
    label_start = last_bsl_param_col + 2
    prod_letter = _col_letter(mid_start + 5)   # Product column (= id_col + 1)
    wf_letter = _col_letter(mid_start + 1)     # Wafer ID column
    # H: Product grouping (show only when product changes)
    for r in range(company_rows):
        row = r + 4
        ws.cell(row=row, column=label_start, value=f"=IF({prod_letter}{row}={prod_letter}{row-1},\"\",{prod_letter}{row})")
    # I: Wafer prefix grouping (show only when first 6 chars of Wafer ID change)
    for r in range(company_rows):
        row = r + 4
        ws.cell(row=row, column=label_start+1, value=f"=IF(LEFT({wf_letter}{row},6)=LEFT({wf_letter}{row-1},6),\"\",LEFT({wf_letter}{row},6))")
    # J: Slot grouping (show last 2 chars of Wafer ID only when it changes)
    for r in range(company_rows):
        row = r + 4
        ws.cell(row=row, column=label_start+2, value=f"=IF({wf_letter}{row}={wf_letter}{row-1},\"\",RIGHT({wf_letter}{row},2))")
    _write_formula_col(ws, mid_start, company_df, company_start_col, "Die Seq", company_rows, "Die Seq")
    _write_formula_col(ws, mid_start + 1, company_df, company_start_col, "Wafer ID", company_rows, "Wafer ID")
    _write_formula_col(ws, mid_start + 2, company_df, company_start_col, "FIELD X", company_rows, "FIELD X")
    _write_formula_col(ws, mid_start + 3, company_df, company_start_col, "FIELD Y", company_rows, "FIELD Y")

    # O: ID = WaferID & "_" & FIELD X & "_" & FIELD Y (company side)
    id_col = mid_start + 4
    ws.cell(row=3, column=id_col, value="ID")
    wx = _col_letter(mid_start + 1)  # Wafer ID col
    wy = _col_letter(mid_start + 2)  # FIELD X col
    wz = _col_letter(mid_start + 3)  # FIELD Y col
    for r in range(company_rows):
        row = r + 4
        ws.cell(row=row, column=id_col, value=f"={wx}{row}&\"_\"&{wy}{row}&\"_\"&{wz}{row}")

    # Product
    wpd = _col_letter(product_col)
    ws.cell(row=3, column=id_col + 1, value=f"Product")
    for r in range(company_rows):
        row = r + 4
        ws.cell(row=row, column=id_col+1, value=f"={wpd}{row}")

    # Raw Company Data
    raw_data_col = id_col+2
    for idx, p in enumerate(selected_params):
        _write_formula_col(ws, raw_data_col+idx, company_df, company_start_col, f"{p}", company_rows, f"{p}")

    # Calibrated Company Data
    calibrated_data_col = raw_data_col + n
    for idx in range(n):
        _slope_str = (f"=slope({_col_letter(raw_data_col + 2 * n + idx)}4:{_col_letter(raw_data_col + 2 * n + idx)}"
                      f"{company_rows+3},{_col_letter(raw_data_col + idx)}4:{_col_letter(raw_data_col + idx)}{company_rows+3})")
        _intercept_str = (f"=intercept({_col_letter(raw_data_col + 2 * n + idx)}4:{_col_letter(raw_data_col + 2 * n + idx)}"
                      f"{company_rows+3},{_col_letter(raw_data_col + idx)}4:{_col_letter(raw_data_col + idx)}{company_rows+3})")
        # print(_slope_str)
        ws.cell(row=1, column=raw_data_col + idx,
                value=_slope_str)
        ws.cell(row=2, column=raw_data_col + idx,
                value=_intercept_str)
    for idx, p in enumerate(selected_params):
        _write_calibrated_data_col(ws, calibrated_data_col+idx, company_rows, p, n)

    # Ref Data
    ref_data_col = calibrated_data_col + n
    for idx, p in enumerate(selected_params):
        _write_ref_data_col(ws, ref_data_col+idx, company_rows, bsl_rows, p, n, idx, id_col)

    # Bias Data
    bias_data_col = ref_data_col + n
    for idx, p in enumerate(selected_params):
        _write_bias_data_col(ws, bias_data_col+idx, company_rows, p, n)
        # ws, target_col, nrows_compay, header, numparam, row_base = 4

    # old Lib Bias Data
    old_bias_data_col = bias_data_col + n
    for idx, p in enumerate(selected_params):
        _write_old_bias_data_col(ws, old_bias_data_col+idx, company_rows, p)

    for r in range(company_rows):
        row = r + 4
        ws.cell(row=row, column=product_col, value="unknown")
        cd = company_df.iloc[r]
        for c_idx, c_name in enumerate(company_df.columns):
            ws.cell(row=row, column=company_start_col + c_idx, value=_safe_val(cd[c_name]))

    wb.save(dst_path)
    return True
